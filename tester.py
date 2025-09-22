import argparse
import concurrent.futures
import os
import sys
import subprocess
import platform
import re
from typing import List, Tuple, Optional, Dict


def _normalize_header(s: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else " " for ch in s).split()


def _choose_column(headers: Dict[int, str], candidates: List[str]) -> Optional[int]:
    # headers: index -> original header text
    # We normalize by tokenizing and matching any token to candidates
    cset = set(candidates)
    for idx, text in headers.items():
        tokens = _normalize_header(str(text))
        # direct token match
        if any(t in cset for t in tokens):
            return idx
        # joined tokens (e.g., ["ip", "address"]) -> "ipaddress"
        joined = "".join(tokens)
        if joined in cset:
            return idx
    return None



def _split_ip_tokens(raw_ip: str) -> List[str]:
    if not raw_ip:
        return []
    # Support comma, semicolon, slash, and whitespace separated IP lists
    parts = [part.strip() for part in re.split(r'[\s,;\/]+', raw_ip) if part.strip()]
    return parts

def load_entries_from_xlsx(path: str,
                           sheet: Optional[str] = None,
                           name_col: Optional[str] = None,
                           ip_col: Optional[str] = None) -> List[Tuple[str, str]]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        print("Error: openpyxl is required to read .xlsx files.")
        print("Install it with: pip install openpyxl")
        raise SystemExit(2)

    if not os.path.exists(path):
        print(f"Error: file not found: {path}")
        raise SystemExit(2)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    headers: Dict[int, str] = {i: (str(v).strip() if v is not None else f"col{i+1}") for i, v in enumerate(header_row)}

    # Resolve columns
    ip_candidates = [
        "ip", "ipaddress", "ipv4", "address"
    ]
    name_candidates = [
        "vm", "vmname", "name", "hostname", "host", "machine", "computer", "server"
    ]

    ip_idx: Optional[int] = None
    name_idx: Optional[int] = None

    if ip_col:
        # Try exact header text match first, then normalized
        for i, h in headers.items():
            if h.lower() == ip_col.lower():
                ip_idx = i
                break
        if ip_idx is None:
            # try normalized join match
            target_norm = "".join(_normalize_header(ip_col))
            for i, h in headers.items():
                if "".join(_normalize_header(h)) == target_norm:
                    ip_idx = i
                    break
    else:
        ip_idx = _choose_column(headers, ip_candidates)

    if name_col:
        for i, h in headers.items():
            if h.lower() == name_col.lower():
                name_idx = i
                break
        if name_idx is None:
            target_norm = "".join(_normalize_header(name_col))
            for i, h in headers.items():
                if "".join(_normalize_header(h)) == target_norm:
                    name_idx = i
                    break
    else:
        name_idx = _choose_column(headers, name_candidates)

    if ip_idx is None:
        print("Error: Could not find IP address column in the header row.")
        print(f"Headers found: {list(headers.values())}")
        print("Specify with --ip-column if needed.")
        raise SystemExit(2)

    # Collect entries
    entries: List[Tuple[str, str]] = []
    for row in rows[1:]:
        # Ensure row has enough columns
        if row is None:
            continue
        ip_val = row[ip_idx] if ip_idx < len(row) else None
        if not ip_val:
            continue
        raw_ip = str(ip_val).strip()
        if not raw_ip:
            continue
        ips = _split_ip_tokens(raw_ip)
        if not ips:
            continue
        name = ""
        if name_idx is not None and name_idx < len(row):
            name_val = row[name_idx]
            if name_val:
                name = str(name_val).strip()
        for ip in ips:
            entries.append((name, ip))

    return entries


def ping_once(target: str, timeout_ms: int = 1000) -> bool:
    is_windows = platform.system().lower().startswith("win")
    if is_windows:
        cmd = ["ping", "-n", "1", "-w", str(timeout_ms), target]
        # Windows timeout applies per echo
        run_timeout = max(1.0, (timeout_ms / 1000.0) + 1.0)
    else:
        # On Unix, -W expects seconds (as integer on many systems)
        timeout_s = max(1, int(round(timeout_ms / 1000.0)))
        cmd = ["ping", "-c", "1", "-W", str(timeout_s), target]
        run_timeout = timeout_s + 2

    try:
        res = subprocess.run(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=run_timeout,
        )
        return res.returncode == 0
    except subprocess.TimeoutExpired:
        return False
    except FileNotFoundError:
        print("Error: 'ping' command not found on this system.")
        raise SystemExit(2)


def check_all(entries: List[Tuple[str, str]], workers: int, timeout_ms: int) -> List[Tuple[str, str, bool]]:
    results: List[Tuple[str, str, bool]] = []
    if workers <= 1:
        for name, ip in entries:
            ok = ping_once(ip, timeout_ms)
            results.append((name, ip, ok))
        return results

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {executor.submit(ping_once, ip, timeout_ms): (name, ip) for name, ip in entries}
        for future in concurrent.futures.as_completed(future_map):
            name, ip = future_map[future]
            ok = False
            try:
                ok = future.result()
            except Exception:
                ok = False
            results.append((name, ip, ok))
    # Keep original order: sort by the index in entries
    order = {t: i for i, t in enumerate(entries)}
    results.sort(key=lambda item: order.get((item[0], item[1]), 0))
    return results


def print_summary(results: List[Tuple[str, str, bool]]):
    # Compute column widths
    name_header = "VM Name"
    ip_header = "IP Address"
    status_header = "Status"

    name_w = max(len(name_header), max((len(n) for n, _, _ in results), default=0))
    ip_w = max(len(ip_header), max((len(ip) for _, ip, _ in results), default=0))

    def line(parts: List[Tuple[str, int]]):
        return "  ".join(text.ljust(width) for text, width in parts)

    print(line([(name_header, name_w), (ip_header, ip_w), (status_header, len(status_header))]))
    print(line([("-" * name_w, name_w), ("-" * ip_w, ip_w), ("-" * len(status_header), len(status_header))]))
    for name, ip, ok in results:
        status = "UP" if ok else "DOWN"
        print(line([(name, name_w), (ip, ip_w), (status, len(status_header))]))




def save_results_to_xlsx(results: List[Tuple[str, str, bool]], output_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        print("Warning: skipping Excel export because openpyxl is unavailable.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Ping Results"
    headers = ["VM Name", "IP Address", "Status"]
    ws.append(headers)

    for name, ip, ok in results:
        ws.append([name, ip, "UP" if ok else "DOWN"])

    header_font = Font(color="FFFFFFFF", bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FF305496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    status_up_fill = PatternFill(fill_type="solid", fgColor="FF92D050")
    status_down_fill = PatternFill(fill_type="solid", fgColor="FFFF5B5B")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        name_cell, ip_cell, status_cell = row
        name_cell.alignment = Alignment(horizontal="left")
        ip_cell.alignment = Alignment(horizontal="left")
        status_text = str(status_cell.value or "").upper()
        status_cell.value = status_text
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.font = Font(bold=True)
        status_cell.fill = status_up_fill if status_text == "UP" else status_down_fill

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[column_letter].width = max(12, max_len + 2)

    ws.auto_filter.ref = f"A1:C{ws.max_row}"

    try:
        wb.save(output_path)
        print(f"Saved results to '{output_path}'.")
    except Exception as exc:
        print(f"Warning: failed to save results to '{output_path}': {exc}")
def parse_args(argv: List[str]):
    p = argparse.ArgumentParser(description="Ping VMs listed in an Excel sheet (.xlsx)")
    p.add_argument("excel", nargs="?", default="ip.xlsx", help="Path to the Excel file (default: ip.xlsx)")
    p.add_argument("--sheet", help="Worksheet name (default: active sheet)")
    p.add_argument("--name-column", dest="name_col", help="Header name for VM name column (auto-detect if omitted)")
    p.add_argument("--ip-column", dest="ip_col", help="Header name for IP address column (auto-detect if omitted)")
    p.add_argument("--timeout", type=int, default=1000, help="Ping timeout per host in ms (default: 1000)")
    p.add_argument("--workers", type=int, default=16, help="Concurrent workers (default: 16; set 1 for sequential)")
    p.add_argument("--output", dest="output_excel", help="Path to write the results workbook (default: <input>_results.xlsx)")
    return p.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)
    entries = load_entries_from_xlsx(args.excel, args.sheet, args.name_col, args.ip_col)
    if not entries:
        print("No entries found in the Excel file.")
        return 1

    print(f"Loaded {len(entries)} entries from '{args.excel}'. Pinging...")
    results = check_all(entries, workers=max(1, args.workers), timeout_ms=max(1, args.timeout))
    print_summary(results)
    output_path = args.output_excel
    if not output_path:
        excel_abspath = os.path.abspath(args.excel)
        excel_dir = os.path.dirname(excel_abspath) or os.getcwd()
        base_name = os.path.splitext(os.path.basename(excel_abspath))[0] or "ping_results"
        output_path = os.path.join(excel_dir, f"{base_name}_results.xlsx")
    save_results_to_xlsx(results, output_path)


    # Exit code 0 if all up; 1 otherwise
    return 0 if all(ok for _, _, ok in results) else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
